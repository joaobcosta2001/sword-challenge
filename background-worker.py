import pika
import json
import time
from openpyxl import load_workbook, Workbook
import threading

# RabbitMQ connection parameters
RABBITMQ_HOST = "rabbitmq"  # Update this if needed
QUEUE_NAME = "recommendations_queue"


print("Starting background worker...")


def process_message(body):
    """
    Process the message received from RabbitMQ.
    This function can log the recommendation, simulate sending notifications, etc.
    """



    # Decode the message
    message = json.loads(body)

    # LOGGING TO A DB
    # The backend is already doing that, so no need to log it here

    # SENDING TO EMAIL OR SMS
    # Not going to do that because i do not have an email address to send to, but if I were to do it I would send this to the AWS API GATWEWAY, then to AWS LAMBDA, and then to SNS, easy :)

    # Define the Excel file path
    excel_file = "/app/data/event_log.xlsx"

    try:
        # Try to load the existing workbook
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        # Add headers to the new file
        sheet.append([ "Timestamp", "Recommendation ID", "Patient ID","Recommendation",])

    # Append the event data to the Excel file
    sheet.append([
        message.get("timestamp"),
        message.get("recommendation_id"),
        message.get("patient_id"),
        message.get("recommendation")
    ])

    # Save the workbook
    workbook.save(excel_file)
    print(f"Event appended to {excel_file}")

    

def callback(ch, method, properties, body):
    """
    Callback function to handle messages from RabbitMQ.
    """
    print(f"Received message: {body}")
    process_message(body)
    ch.basic_ack(delivery_tag=method.delivery_tag)  # Acknowledge the message

# Establish RabbitMQ connection
for _ in range(10):  # Retry up to 5 times
    try:
        connection = pika.BlockingConnection(pika.ConnectionParameters(host=RABBITMQ_HOST))
        channel = connection.channel()
        channel.queue_declare(queue=QUEUE_NAME)
        break
    except pika.exceptions.AMQPConnectionError:
        print("Failed to connect to RabbitMQ. Retrying in 5 seconds...")
        time.sleep(5)
else:
    raise Exception("Could not connect to RabbitMQ after multiple attempts.")

# Start consuming messages
channel.basic_consume(queue=QUEUE_NAME, on_message_callback=callback)


def start_consuming():
    while True:
        try:
            print("Starting to consume messages...")
            channel.start_consuming()
        except Exception as e:
            print(f"Error in consumer thread: {e}. Retrying in 5 seconds...")
            time.sleep(5)

def heartbeat():
    while True:
        print("Worker is alive...")
        time.sleep(10)

# Start a heartbeat thread
heartbeat_thread = threading.Thread(target=heartbeat, daemon=True)
heartbeat_thread.start()

# Start the consuming loop in a separate thread
consumer_thread = threading.Thread(target=start_consuming)
consumer_thread.start()

print("Background worker is running and ready to consume messages.")